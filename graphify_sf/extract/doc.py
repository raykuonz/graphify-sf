"""Document extractor for graphify-sf.

Handles non-Salesforce reference files:
  - Markdown / plain text / RST / HTML  → headings as sub-nodes
  - PDF                                  → text extraction via pypdf
  - Images                               → metadata node only

SF component name mention detection creates INFERRED references edges
back to SF nodes, resolved in the cross-reference pass.
"""
from __future__ import annotations

import hashlib
import re
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# ID helpers
# ---------------------------------------------------------------------------

def _doc_id(path: Path, root: Path | None = None) -> str:
    """Stable node ID for a document file."""
    key = str(path.resolve()) if root is None else str(path)
    h = hashlib.sha256(key.encode()).hexdigest()[:8]
    stem = re.sub(r"[^a-z0-9]", "_", path.stem.lower())
    return f"doc_{stem}_{h}"


def _heading_id(doc_id: str, text: str, lineno: int) -> str:
    slug = re.sub(r"[^a-z0-9]", "_", text.lower())[:40].strip("_")
    return f"{doc_id}_h_{slug}_{lineno}"


def _xlsx_sheet_id(stem: str, sheet_name: str) -> str:
    slug = re.sub(r"[^a-z0-9]", "_", f"{stem}_{sheet_name}".lower())
    return f"xlsx_{slug}"


def _xlsx_table_id(stem: str, sheet_name: str, table_name: str) -> str:
    slug = re.sub(r"[^a-z0-9]", "_", f"{stem}_{sheet_name}_{table_name}".lower())
    return f"xlsx_{slug}"


def _xlsx_col_id(stem: str, sheet_name: str, table_name: str, col_name: str) -> str:
    slug = re.sub(r"[^a-z0-9]", "_", f"{stem}_{sheet_name}_{table_name}_{col_name}".lower())
    return f"xlsx_{slug}"


# ---------------------------------------------------------------------------
# Heading extraction
# ---------------------------------------------------------------------------

_HEADING_RE = re.compile(r"^(#{1,3})\s+(.+)$", re.MULTILINE)


def _extract_headings(text: str, doc_id: str, source_file: str) -> tuple[list[dict], list[dict]]:
    """Extract H1/H2/H3 headings as sub-nodes with contains edges."""
    nodes: list[dict] = []
    edges: list[dict] = []
    for lineno, line in enumerate(text.splitlines(), start=1):
        m = _HEADING_RE.match(line)
        if m:
            level = len(m.group(1))
            heading_text = m.group(2).strip()
            h_id = _heading_id(doc_id, heading_text, lineno)
            nodes.append({
                "id": h_id,
                "label": heading_text,
                "file_type": "document",
                "sf_type": None,
                "source_file": source_file,
                "source_location": f"L{lineno}",
                "heading_level": level,
            })
            edges.append({
                "source": doc_id,
                "target": h_id,
                "relation": "contains",
                "confidence": "EXTRACTED",
                "confidence_score": 1.0,
                "source_file": source_file,
                "source_location": f"L{lineno}",
                "weight": 0.5,
            })
    return nodes, edges


# ---------------------------------------------------------------------------
# SF mention detection
# ---------------------------------------------------------------------------

def _sf_mention_edges(text: str, doc_id: str, source_file: str) -> list[dict]:
    """Scan text for SF API name patterns (e.g. AccountService, Account__c).

    Emits INFERRED reference edges — targets are resolved in the cross-reference
    pass once all SF nodes are known.
    """
    edges: list[dict] = []
    # SF API names: PascalCase identifiers or names ending in __c / __r / __mdt etc.
    sf_pattern = re.compile(
        r"\b([A-Z][A-Za-z0-9]*(?:__[a-z]{1,6})?)\b"
    )
    seen: set[str] = set()
    for m in sf_pattern.finditer(text):
        name = m.group(1)
        if name in seen or len(name) < 3:
            continue
        seen.add(name)
        # Candidate SF node IDs to try at cross-reference time
        # We store the raw mention label in _mention_label for resolution
        edges.append({
            "source": doc_id,
            "target": f"__mention__{name}",   # placeholder resolved in pass 2
            "_mention_label": name,
            "relation": "references",
            "confidence": "INFERRED",
            "confidence_score": 0.6,
            "source_file": source_file,
            "source_location": None,
            "weight": 0.5,
        })
    return edges


# ---------------------------------------------------------------------------
# xlsx structure extraction
# ---------------------------------------------------------------------------

def xlsx_extract_structure(path: Path) -> dict:
    """Extract structural nodes (file → sheet → named_table → column) from an .xlsx.

    Returns {nodes, edges} compatible with the graphify-sf extract pipeline.
    """
    str_path = str(path)
    stem = re.sub(r"[^a-z0-9]", "_", path.stem.lower())
    file_id = _doc_id(path)
    nodes: list[dict] = [{
        "id": file_id,
        "label": path.name,
        "file_type": "document",
        "sf_type": None,
        "source_file": str_path,
        "source_location": None,
    }]
    edges: list[dict] = []
    seen: set[str] = {file_id}

    def _add_node(nid: str, label: str) -> None:
        if nid not in seen:
            seen.add(nid)
            nodes.append({
                "id": nid, "label": label,
                "file_type": "document", "sf_type": None,
                "source_file": str_path, "source_location": None,
            })

    def _add_edge(src: str, tgt: str, relation: str) -> None:
        edges.append({
            "source": src, "target": tgt, "relation": relation,
            "confidence": "EXTRACTED", "confidence_score": 1.0,
            "source_file": str_path, "source_location": None, "weight": 0.5,
        })

    try:
        import openpyxl
    except ImportError:
        return {"nodes": nodes, "edges": edges}

    try:
        wb = openpyxl.load_workbook(str_path, read_only=False, data_only=True)
    except Exception as exc:
        print(f"[graphify-sf] WARNING: xlsx_extract_structure failed for {path}: {exc}", file=sys.stderr)
        return {"nodes": nodes, "edges": edges}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_id = _xlsx_sheet_id(stem, sheet_name)
        _add_node(sheet_id, f"{sheet_name} (sheet)")
        _add_edge(file_id, sheet_id, "contains")

        if hasattr(ws, "tables"):
            for tbl in ws.tables.values():
                tbl_id = _xlsx_table_id(stem, sheet_name, tbl.name)
                _add_node(tbl_id, tbl.name)
                _add_edge(sheet_id, tbl_id, "contains")
                if tbl.ref:
                    try:
                        from openpyxl.utils import range_boundaries
                        min_col, min_row, max_col, _ = range_boundaries(tbl.ref)
                        header_rows = list(ws.iter_rows(
                            min_row=min_row, max_row=min_row,
                            min_col=min_col, max_col=max_col,
                            values_only=True,
                        ))
                        if header_rows:
                            for col_name in header_rows[0]:
                                if col_name:
                                    col_id = _xlsx_col_id(stem, sheet_name, tbl.name, str(col_name))
                                    _add_node(col_id, str(col_name))
                                    _add_edge(tbl_id, col_id, "contains")
                    except Exception:
                        pass
        else:
            # Fallback: use first non-empty row as column headers
            for row in ws.iter_rows(max_row=1, values_only=True):
                for cell in row:
                    if cell:
                        col_id = _xlsx_col_id(stem, sheet_name, "data", str(cell))
                        _add_node(col_id, str(cell))
                        _add_edge(sheet_id, col_id, "contains")
                break

    try:
        wb.close()
    except Exception:
        pass

    return {"nodes": nodes, "edges": edges}


# ---------------------------------------------------------------------------
# Main extractors
# ---------------------------------------------------------------------------

def extract_document(path: Path) -> dict:
    """Extract a text document (.md, .txt, .rst, .html) into graph nodes/edges."""
    str_path = str(path)
    doc_id = _doc_id(path)
    nodes: list[dict] = [{
        "id": doc_id,
        "label": path.name,
        "file_type": "document",
        "sf_type": None,
        "source_file": str_path,
        "source_location": None,
    }]
    edges: list[dict] = []

    try:
        text = path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return {"nodes": nodes, "edges": edges}

    h_nodes, h_edges = _extract_headings(text, doc_id, str_path)
    nodes.extend(h_nodes)
    edges.extend(h_edges)
    edges.extend(_sf_mention_edges(text, doc_id, str_path))

    return {"nodes": nodes, "edges": edges}


def extract_paper(path: Path) -> dict:
    """Extract a PDF document into a graph node."""
    from graphify_sf.detect import extract_pdf_text
    str_path = str(path)
    doc_id = _doc_id(path)
    nodes: list[dict] = [{
        "id": doc_id,
        "label": path.name,
        "file_type": "paper",
        "sf_type": None,
        "source_file": str_path,
        "source_location": None,
    }]
    edges: list[dict] = []

    text = extract_pdf_text(path)
    if text:
        edges.extend(_sf_mention_edges(text, doc_id, str_path))

    return {"nodes": nodes, "edges": edges}


def extract_image(path: Path) -> dict:
    """Create a metadata-only node for an image file."""
    str_path = str(path)
    doc_id = _doc_id(path)
    return {
        "nodes": [{
            "id": doc_id,
            "label": path.name,
            "file_type": "image",
            "sf_type": None,
            "source_file": str_path,
            "source_location": None,
        }],
        "edges": [],
    }


def extract_doc_file(path: Path) -> dict:
    """Dispatch to the right document extractor based on file extension."""
    ext = path.suffix.lower()
    if ext in {".md", ".mdx", ".txt", ".rst", ".html", ".htm"}:
        return extract_document(path)
    if ext == ".pdf":
        return extract_paper(path)
    if ext in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"}:
        return extract_image(path)
    # .xlsx sidecar (already converted to .md by detect()) — treat as document
    if ext == ".xlsx":
        return xlsx_extract_structure(path)
    # Fallback for converted sidecars (.md from .docx/.xlsx)
    return extract_document(path)
